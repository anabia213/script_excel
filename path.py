import time
import openpyxl
import os
import rrdtool

def create_rrd(rrd_filename):
    # Definição das fontes de dados e arquivos de arquivamento round-robin
    data_sources = [
        "DS:Timestamp:GAUGE:5:U:U",
        "DS:Point:GAUGE:5:U:U",
        "DS:Downlink_Throughp:GAUGE:5:U:U",
        "DS:Uplink_Throughp:GAUGE:5:U:U",
        "DS:Ping_Drop_Rate_Avg:GAUGE:5:U:U",
        "DS:Ping_Latency_Ms_Avg:GAUGE:5:U:U",
        "DS:Obstruction_Percent_Time:GAUGE:5:U:U",
        "DS:Signal_Quality:GAUGE:5:U:U"
    ]
    round_robin_archives = [
        "RRA:AVERAGE:0.5:1:100",
        "RRA:AVERAGE:0.5:6:600"
    ]
    
    # Criação do arquivo RRD
    rrdtool.create(rrd_filename, "--step", "5", *data_sources, *round_robin_archives)

def update_rrd(spreadsheet_path):
    # Atualiza o banco de dados RRD com os novos valores da planilha
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook.active
    for col_index in range(1, sheet.max_column + 1):
        column = sheet.cell(row=1, column=col_index).value
        value = sheet.cell(row=sheet.max_row, column=col_index).value
        rrdtool.update('planilha.rrd', 'N:{}'.format(value))

def create_graph(rrdtool_path, data, output_path):
    # Criação do gráfico com base nos dados do arquivo RRD
    command = [rrdtool_path, 'graph', output_path]
    for column, values in data.items():
        command.extend(['DEF:{0}=planilha.rrd:{0}:AVERAGE'.format(column),
                        'LINE1:{0}#{1}:{0}'.format(column, values['color']),
                        'GPRINT:{0}:LAST: {0} \: %6.2lf%s'.format(column)])
    os.system(' '.join(command))

def monitor_spreadsheet(spreadsheet_path, rrdtool_path):
    # Monitoramento contínuo da planilha
    workbook = openpyxl.load_workbook(spreadsheet_path)
    sheet = workbook.active
    previous_rows = sheet.max_row

    while True:
        time.sleep(1)  # Aguarda 1 segundo antes de verificar novamente

        # Verifica se houve atualização na planilha
        if sheet.max_row > previous_rows:
            # Se houve, atualiza o banco de dados RRD com os novos valores
            update_rrd(spreadsheet_path)
            # Atualiza o número de linhas anteriores
            previous_rows = sheet.max_row

            # Define os dados para criar o gráfico
            data = {
                'Point': {'color': 'FF0000'},  # Cor do gráfico para 'Point'
                'Downlink_Throughp': {'color': '00FF00'},  # Cor do gráfico para 'Downlink_Throughp'
                'Uplink_Throughp': {'color': '0000FF'},  # Cor do gráfico para 'Uplink_Throughp'
                'Ping_Drop_Rate_Avg': {'color': 'FFFF00'},  # Cor do gráfico para 'Ping_Drop_Rate_Avg'
                'Ping_Latency_Ms_Avg': {'color': '00FFFF'},  # Cor do gráfico para 'Ping_Latency_Ms_Avg'
                'Obstruction_Percent_Time': {'color': 'FF00FF'},  # Cor do gráfico para 'Obstruction_Percent_Time'
                'Signal_Quality': {'color': '000000'}  # Cor do gráfico para 'Signal_Quality'
            }
            # Cria o gráfico com base nos dados do arquivo RRD
            create_graph(rrdtool_path, data, 'output_graph.png')

if __name__ == "__main__":
    spreadsheet_path = '/home/anabia/path_starlink/starlink_ln.xlsx'  # Caminho da planilha Excel
    rrd_filename = 'planilha.rrd'  # Nome do arquivo RRD
    rrdtool_path = '/usr/bin/rrdtool'  # Caminho do rrdtool

    # Verifica se o arquivo RRD já existe, se não, cria-o
    if not os.path.exists(rrd_filename):
        create_rrd(rrd_filename)

    # Inicia o monitoramento da planilha
    monitor_spreadsheet(spreadsheet_path, rrdtool_path)

